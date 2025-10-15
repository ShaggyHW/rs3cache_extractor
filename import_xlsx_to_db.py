#!/usr/bin/env python3
"""
Import rows from an .xlsx workbook into the SQLite database tables.

- Each worksheet name is matched to a table of the same name in the DB.
- Row 1 must contain column headers that correspond to the DB column names.
- Types are coerced using the DB schema (PRAGMA table_info).
- Primary key columns may be omitted or left blank to auto-generate (when allowed).
- Supports dry-run and optional truncation per table.
- For tables with a single-column primary key (e.g., `id`), performs
  INSERT ... ON CONFLICT(id) DO UPDATE SET ... so rows are inserted or
  updated based on the primary key. When the PK is omitted/blank, a
  normal INSERT is performed (allowing auto-generated ids where defined).

Usage:
  python import_xlsx_to_db.py --xlsx ImportSpreadSheet.xlsx [--db worldReachableTiles.db]
  python import_xlsx_to_db.py --xlsx ImportSpreadSheet.xlsx --dry-run
  python import_xlsx_to_db.py --xlsx ImportSpreadSheet.xlsx --truncate requirements door_nodes
  python import_xlsx_to_db.py --xlsx "https://docs.google.com/spreadsheets/d/<sheet_id>/edit?gid=<gid>#gid=<gid>"

Notes:
- Requires: openpyxl (pip install openpyxl)
- Recognizes and validates constrained values like 'next_node_type' and door 'direction'.
- For requirements.comparison supports: =, !=, <, <=, >, >=
"""

from __future__ import annotations

import argparse
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple
from urllib.parse import urlparse, parse_qs
from urllib.request import urlopen, Request
import tempfile
import os

# Lazy import openpyxl with a friendly error if missing
try:
    from openpyxl import load_workbook  # type: ignore
except Exception as e:  # pragma: no cover
    load_workbook = None  # type: ignore


ALLOWED_NEXT_NODE_TYPES = {"object", "npc", "ifslot", "door", "lodestone", "item"}
ALLOWED_DOOR_DIRECTIONS = {"IN", "OUT"}
ALLOWED_REQUIREMENT_COMPARISONS = {"=", "!=", "<", "<=", ">", ">="}


def is_google_sheets_url(s: str) -> bool:
    try:
        p = urlparse(s)
    except Exception:
        return False
    return (
        p.scheme in {"http", "https"}
        and "docs.google.com" in (p.netloc or "")
        and "/spreadsheets/" in (p.path or "")
    )


def build_gsheet_export_url(doc_url: str) -> str:
    p = urlparse(doc_url)
    # Typical path: /spreadsheets/d/<sheet_id>/edit
    parts = [seg for seg in (p.path or "").split("/") if seg]
    sheet_id: Optional[str] = None
    for i, seg in enumerate(parts):
        if seg == "d" and i + 1 < len(parts):
            sheet_id = parts[i + 1]
            break
    if not sheet_id:
        raise ValueError("Unable to parse Google Sheets ID from URL")

    query = parse_qs(p.query or "")
    gid = query.get("gid", [None])[0]
    base = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    if gid:
        return f"{base}&gid={gid}"
    return base


def download_google_sheet_as_xlsx(doc_url: str) -> Path:
    export_url = build_gsheet_export_url(doc_url)
    req = Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req) as resp:
        data = resp.read()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        tmp.write(data)
        tmp.flush()
    finally:
        tmp.close()
    return Path(tmp.name)


@dataclass
class Column:
    name: str
    decl_type: str  # as returned by PRAGMA table_info (e.g., 'INTEGER', 'TEXT')
    notnull: bool
    pk: bool

    @property
    def is_integer(self) -> bool:
        return "INT" in self.decl_type.upper()

    @property
    def is_text(self) -> bool:
        return "CHAR" in self.decl_type.upper() or "CLOB" in self.decl_type.upper() or "TEXT" in self.decl_type.upper()


@dataclass
class Table:
    name: str
    columns: Dict[str, Column]  # keyed by lowercased name for case-insensitive matching


def get_connection(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def fetch_existing_tables(conn: sqlite3.Connection) -> Dict[str, Table]:
    tables: Dict[str, Table] = {}
    cur = conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
    for row in cur:
        tname = row["name"]
        # Skip SQLite internal tables
        if tname.startswith("sqlite_"):
            continue
        # Introspect columns
        cols_cur = conn.execute(f"PRAGMA table_info('{tname}')")
        cols: Dict[str, Column] = {}
        for c in cols_cur:
            col = Column(
                name=c[1],
                decl_type=c[2] or "",
                notnull=bool(c[3]),
                pk=bool(c[5]),
            )
            cols[c[1].lower()] = col
        tables[tname.lower()] = Table(name=tname, columns=cols)
    return tables


def normalize_header(header: Any) -> Optional[str]:
    if header is None:
        return None
    name = str(header).strip()
    if not name:
        return None
    return name


def coerce_value(raw: Any, col: Column, table_name: str) -> Any:
    if raw is None:
        return None
    # If cell has been read as string, normalize whitespace
    if isinstance(raw, str):
        val = raw.strip()
        if val == "":
            return None
        # Booleans encoded as strings
        if col.is_integer:
            low = val.lower()
            if low in ("true", "yes", "y", "on"):  # bool-like
                return 1
            if low in ("false", "no", "n", "off"):
                return 0
            # Try int cast then float->int
            try:
                return int(val)
            except ValueError:
                try:
                    f = float(val)
                    if f.is_integer():
                        return int(f)
                    return int(round(f))
                except ValueError:
                    pass
        return val

    # Numeric types from openpyxl
    if isinstance(raw, bool):
        if col.is_integer:
            return 1 if raw else 0
        return str(raw)
    if isinstance(raw, (int,)):
        if col.is_integer:
            return int(raw)
        return str(raw)
    if isinstance(raw, float):
        if col.is_integer:
            return int(raw) if raw.is_integer() else int(round(raw))
        return str(raw)

    return raw


def validate_specials(table_name: str, row: Dict[str, Any]) -> None:
    t = table_name.lower()
    if t == "door_nodes":
        if "direction" in row and row["direction"] is not None:
            row["direction"] = str(row["direction"]).strip().upper()
            if row["direction"] not in ALLOWED_DOOR_DIRECTIONS:
                raise ValueError(f"Invalid door_nodes.direction: {row['direction']} (allowed {sorted(ALLOWED_DOOR_DIRECTIONS)})")
    if "next_node_type" in row and row["next_node_type"] is not None:
        row["next_node_type"] = str(row["next_node_type"]).strip().lower()
        if row["next_node_type"] not in ALLOWED_NEXT_NODE_TYPES:
            raise ValueError(
                f"Invalid next_node_type: {row['next_node_type']} (allowed {sorted(ALLOWED_NEXT_NODE_TYPES)})"
            )
    if t == "requirements":
        if "comparison" in row and row["comparison"] is not None:
            cmpv = str(row["comparison"]).strip()
            if cmpv not in ALLOWED_REQUIREMENT_COMPARISONS:
                raise ValueError(
                    f"Invalid requirements.comparison: {cmpv} (allowed {sorted(ALLOWED_REQUIREMENT_COMPARISONS)})"
                )


def build_insert_sql(table: Table, row: Dict[str, Any]) -> Tuple[str, Tuple[Any, ...]]:
    # Filter to known columns and decide whether to include PK if provided
    cols: List[str] = []
    vals: List[Any] = []
    for key, val in row.items():
        col = table.columns.get(key.lower())
        if not col:
            continue
        if col.pk and (val is None or val == ""):
            # Skip PK when empty to allow auto generation
            continue
        cols.append(col.name)
        vals.append(val)
    if not cols:
        raise ValueError("No valid columns to insert after filtering")

    # Determine single-column primary key, if present
    pk_names = [c.name for c in table.columns.values() if c.pk]
    pk_name: Optional[str] = pk_names[0] if len(pk_names) == 1 else None

    placeholders = ",".join(["?"] * len(cols))
    base_insert = f"INSERT INTO {table.name} ({', '.join(cols)}) VALUES ({placeholders})"

    # If PK exists and provided in the row, use UPSERT to update on conflict
    if pk_name and pk_name in cols:
        # Build update set for non-PK columns included in this row
        assignments = [f"{c}=excluded.{c}" for c in cols if c != pk_name]
        if assignments:
            sql = base_insert + f" ON CONFLICT({pk_name}) DO UPDATE SET " + ", ".join(assignments)
        else:
            # Only PK provided; use OR IGNORE to avoid errors
            sql = base_insert.replace("INSERT ", "INSERT OR IGNORE ")
        return sql, tuple(vals)

    # Fallback: plain INSERT (no PK provided or composite PK unsupported here)
    return base_insert, tuple(vals)


def read_worksheet(ws, table: Table) -> List[Dict[str, Any]]:
    # Determine headers from the first row
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return []
    headers: List[Optional[str]] = [normalize_header(h) for h in headers_raw]
    # Map headers to DB columns by case-insensitive name
    header_to_colname: List[Optional[str]] = []
    for h in headers:
        if h is None:
            header_to_colname.append(None)
        else:
            key = h.lower()
            if key in table.columns:
                header_to_colname.append(table.columns[key].name)
            else:
                header_to_colname.append(None)  # unknown header ignored
    out_rows: List[Dict[str, Any]] = []
    for data_row in rows_iter:
        if data_row is None:
            continue
        # Build row dict mapping column name -> value
        row_map: Dict[str, Any] = {}
        empty = True
        for idx, raw in enumerate(data_row):
            if idx >= len(header_to_colname):
                break
            colname = header_to_colname[idx]
            if colname is None:
                continue
            col = table.columns[colname.lower()]
            val = coerce_value(raw, col, table.name)
            if val is not None and val != "":
                empty = False
            row_map[colname] = val
        if empty:
            continue  # skip completely empty lines
        validate_specials(table.name, row_map)
        out_rows.append(row_map)
    return out_rows


def import_workbook(xlsx_path: Path, db_path: Path, truncate: Sequence[str], dry_run: bool, only_sheets: Optional[Sequence[str]]) -> None:
    if load_workbook is None:
        raise SystemExit("openpyxl is required. Install with: pip install openpyxl")

    wb = load_workbook(filename=str(xlsx_path), data_only=True)

    with get_connection(db_path) as conn:
        tables = fetch_existing_tables(conn)
        truncate_set = {t.lower() for t in truncate}
        only_set = {s.lower() for s in only_sheets} if only_sheets else None

        # Validate requested truncations and sheets
        for t in truncate_set:
            if t not in tables:
                raise SystemExit(f"--truncate table not found in DB: {t}")
        if only_set:
            for s in only_set:
                if s not in tables:
                    raise SystemExit(f"Requested sheet/table not found in DB: {s}")

        # Begin transaction
        conn.execute("BEGIN")
        try:
            # Truncate if requested
            for t in truncate_set:
                print(f"Truncating table: {tables[t].name}")
                if not dry_run:
                    conn.execute(f"DELETE FROM {tables[t].name}")

            total_inserted = 0
            # Process 'requirements' first to satisfy foreign key references
            sheets = list(wb.worksheets)
            sheets.sort(key=lambda ws: 0 if ws.title.lower() == "requirements" else 1)
            for ws in sheets:
                sheet_name = ws.title
                sheet_key = sheet_name.lower()
                if only_set and sheet_key not in only_set:
                    continue
                if sheet_key not in tables:
                    print(f"Skipping worksheet '{sheet_name}' (no matching table in DB)")
                    continue
                table = tables[sheet_key]
                print(f"Processing worksheet '{sheet_name}' -> table '{table.name}'")
                rows = read_worksheet(ws, table)
                print(f"  Prepared {len(rows)} row(s)")
                sheet_preview = 0
                for r in rows:
                    sql, params = build_insert_sql(table, r)
                    if dry_run:
                        # Show a preview for the first few rows per sheet
                        if sheet_preview < 5:
                            print(f"  SQL: {sql}\n  Params: {params}")
                            sheet_preview += 1
                    else:
                        conn.execute(sql, params)
                    total_inserted += 1
            if dry_run:
                print(f"Dry-run complete. Rows that would be inserted: {total_inserted}")
                conn.execute("ROLLBACK")
            else:
                conn.execute("COMMIT")
                print(f"Import complete. Rows inserted: {total_inserted}")
        except Exception as e:
            conn.execute("ROLLBACK")
            raise


def main() -> None:
    parser = argparse.ArgumentParser(description="Import .xlsx worksheets into SQLite tables")
    parser.add_argument("--xlsx", required=True, type=str, help="Path to .xlsx file or Google Sheets URL")
    parser.add_argument("--db", default=Path("worldReachableTiles.db"), type=Path, help="Path to SQLite DB (default: worldReachableTiles.db)")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate only; do not modify the DB")
    parser.add_argument("--truncate", nargs="*", default=[], help="Tables to DELETE FROM before inserting (by name, e.g., requirements door_nodes)")
    parser.add_argument("--sheets", nargs="*", default=None, help="Only import the specified sheets/tables")
    args = parser.parse_args()

    # Determine local path vs Google Sheets URL
    downloaded_temp = False
    if is_google_sheets_url(args.xlsx):
        print("Downloading Google Sheet as .xlsx ...")
        try:
            xlsx_path = download_google_sheet_as_xlsx(args.xlsx)
            downloaded_temp = True
        except Exception as e:
            raise SystemExit(f"Failed to download Google Sheet: {e}")
    else:
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.exists():
            raise SystemExit(f"XLSX file not found: {xlsx_path}")
    if not args.db.exists():
        raise SystemExit(f"SQLite DB not found: {args.db}")

    try:
        import_workbook(xlsx_path, args.db, args.truncate, args.dry_run, args.sheets)
    finally:
        if downloaded_temp:
            try:
                os.unlink(xlsx_path)
            except Exception:
                pass


if __name__ == "__main__":
    main()
